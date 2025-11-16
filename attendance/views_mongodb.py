from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
from bson import ObjectId
import csv
import json

from .mongodb_utils import MongoDBManager
from .utils_photo import generate_initials_photo

# Initialize MongoDB managers
devotees_db = MongoDBManager('devotees')
sabhas_db = MongoDBManager('sabhas')
attendance_db = MongoDBManager('attendance_records')

@csrf_exempt
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        messages.error(request, 'Invalid credentials')
    return render(request, 'registration/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):
    total_devotees = devotees_db.count()
    recent_sabhas_raw = sabhas_db.find(sort=[('date', -1)], limit=5)
    recent_sabhas = []
    for sabha in recent_sabhas_raw:
        sabha['id'] = str(sabha['_id'])
        sabha['get_sabha_type_display'] = sabha.get('sabha_type', '').title() + ' Sabha'
        recent_sabhas.append(sabha)
    
    total_attendance_records = attendance_db.count()
    present_records = attendance_db.count({'status': 'present'})
    attendance_rate = round((present_records / total_attendance_records * 100) if total_attendance_records > 0 else 0)
    
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    this_week_sabhas = sabhas_db.count({
        'date': {'$gte': week_start.isoformat(), '$lte': week_end.isoformat()}
    })
    
    context = {
        'total_devotees': total_devotees,
        'recent_sabhas': recent_sabhas,
        'attendance_rate': attendance_rate,
        'this_week_sabhas': this_week_sabhas,
    }
    return render(request, 'attendance/dashboard.html', context)

@login_required
def devotee_list(request):
    search_query = request.GET.get('search', '')
    search_type = request.GET.get('search_type', 'id')
    page = int(request.GET.get('page', 1))
    per_page = 20
    
    query = {}
    if search_query:
        if search_type == 'id':
            # Search both as integer and string to handle mixed data types
            if search_query.isdigit():
                query = {'$or': [
                    {'devotee_id': int(search_query)},
                    {'devotee_id': str(search_query)}
                ]}
            else:
                query = {'devotee_id': {'$regex': str(search_query), '$options': 'i'}}
        elif search_type == 'phone':
            query = {'contact_number': {'$regex': search_query}}
        elif search_type == 'name':
            query = {'name': {'$regex': search_query, '$options': 'i'}}
        elif search_type == 'type':
            query = {'devotee_type': {'$regex': search_query, '$options': 'i'}}
    
    total_count = devotees_db.count(query)
    skip = (page - 1) * per_page
    
    # Get all matching documents and sort by devotee_id as integer
    all_devotees = list(devotees_db.find(query))
    all_devotees.sort(key=lambda x: int(x.get('devotee_id', 0)) if str(x.get('devotee_id', '')).isdigit() else 0)
    devotees_raw = all_devotees[skip:skip + per_page]
    devotees = []
    for devotee in devotees_raw:
        devotee['id'] = str(devotee['_id'])
        devotee['get_sabha_type_display'] = devotee.get('sabha_type', '').title()
        if not devotee.get('photo_url'):
            devotee['photo_url'] = generate_initials_photo(devotee['name'])
        devotees.append(devotee)
    
    # Pagination info
    total_pages = (total_count + per_page - 1) // per_page
    has_previous = page > 1
    has_next = page < total_pages
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        devotees_data = []
        for d in devotees_raw:
            photo_url = d.get('photo_url', '')
            if not photo_url:
                photo_url = generate_initials_photo(d['name'])
            devotees_data.append({
                'id': str(d['_id']),
                'devotee_id': d.get('devotee_id', ''),
                'name': d['name'],
                'contact_number': d['contact_number'],
                'sabha_type_display': d['sabha_type'].title(),
                'devotee_type': d.get('devotee_type', ''),
                'contact_number': d['contact_number'],
                'join_date': d['join_date'],
                'photo_url': photo_url
            })
        
        return JsonResponse({
            'devotees': devotees_data,
            'total_count': total_count,
            'current_page': page,
            'total_pages': total_pages,
            'has_previous': has_previous,
            'has_next': has_next
        })
    
    # Create pagination object for template
    class PaginationObj:
        def __init__(self, items, page, total_pages, has_previous, has_next, total_count):
            self.object_list = items
            self.number = page
            self.paginator = type('Paginator', (), {
                'num_pages': total_pages,
                'count': total_count
            })()
            self.has_previous = lambda: has_previous
            self.has_next = lambda: has_next
            self.previous_page_number = lambda: page - 1 if has_previous else None
            self.next_page_number = lambda: page + 1 if has_next else None
        
        def __iter__(self):
            return iter(self.object_list)
    
    page_obj = PaginationObj(devotees, page, total_pages, has_previous, has_next, total_count)
    
    return render(request, 'attendance/devotee_list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
        'search_type': search_type,
        'total_count': total_count
    })

@login_required
def devotee_add(request):
    if request.method == 'POST':
        from .dropbox_utils import upload_devotee_photo
        
        # Auto-generate devotee_id if not provided - MongoDB utils will handle this
        devotee_id = request.POST.get('devotee_id', '').strip()
        if devotee_id and str(devotee_id).isdigit():
            devotee_id = int(devotee_id)
        elif not devotee_id:
            devotee_id = None  # Let MongoDB utils auto-generate
        
        devotee_data = {
            'devotee_type': request.POST.get('devotee_type', 'haribhakt'),
            'name': request.POST.get('name'),
            'contact_number': request.POST.get('contact_number'),
            'date_of_birth': request.POST.get('date_of_birth'),
            'gender': request.POST.get('gender'),
            'age': int(request.POST.get('age', 0)),
            'sabha_type': request.POST.get('sabha_type'),
            'address_line': request.POST.get('address_line'),
            'landmark': request.POST.get('landmark'),
            'zone': request.POST.get('zone'),
            'join_date': request.POST.get('join_date'),
            'photo_url': '',
            'created_at': datetime.now().isoformat()
        }
        
        # Handle photo upload
        if request.POST.get('cropped_photo'):
            # Handle cropped photo from base64
            import base64
            from io import BytesIO
            from django.core.files.uploadedfile import InMemoryUploadedFile
            
            cropped_data = request.POST.get('cropped_photo')
            format, imgstr = cropped_data.split(';base64,')
            ext = format.split('/')[-1]
            
            data = base64.b64decode(imgstr)
            photo_file = InMemoryUploadedFile(
                BytesIO(data), None, f'cropped_photo.{ext}', f'image/{ext}', len(data), None
            )
            
            photo_url = upload_devotee_photo(photo_file, devotee_data)
            if photo_url:
                devotee_data['photo_url'] = photo_url
        elif 'photo' in request.FILES:
            # Handle regular photo upload
            photo_file = request.FILES['photo']
            photo_url = upload_devotee_photo(photo_file, devotee_data)
            if photo_url:
                devotee_data['photo_url'] = photo_url
        
        # Only add devotee_id if provided, otherwise let MongoDB utils auto-generate
        if devotee_id:
            devotee_data['devotee_id'] = devotee_id
        
        result = devotees_db.insert_one(devotee_data)
        
        if result is None:
            messages.error(request, 'Database connection error. Please try again.')
            return render(request, 'attendance/devotee_form.html', {'title': 'Add Devotee', 'today': datetime.now().date().isoformat()})
        
        messages.success(request, 'Devotee added successfully!')
        return redirect('devotee_list')
    
    return render(request, 'attendance/devotee_form.html', {'title': 'Add Devotee', 'today': datetime.now().date().isoformat()})

@login_required
def devotee_detail(request, pk):
    devotee = devotees_db.find_one({'_id': ObjectId(pk)})
    if not devotee:
        messages.error(request, 'Devotee not found')
        return redirect('devotee_list')
    devotee['id'] = str(devotee['_id'])
    devotee['pk'] = str(devotee['_id'])
    devotee['get_sabha_type_display'] = devotee.get('sabha_type', '').title()
    return render(request, 'attendance/devotee_detail.html', {'devotee': devotee})

@login_required
def devotee_edit(request, pk):
    devotee = devotees_db.find_one({'_id': ObjectId(pk)})
    if not devotee:
        messages.error(request, 'Devotee not found')
        return redirect('devotee_list')
    
    if request.method == 'POST':
        from .dropbox_utils import upload_devotee_photo
        
        devotee_id_input = request.POST.get('devotee_id', devotee.get('devotee_id', ''))
        # Ensure devotee_id is stored as integer if it's numeric
        if str(devotee_id_input).isdigit():
            devotee_id_input = int(devotee_id_input)
        
        update_data = {
            'devotee_id': devotee_id_input,
            'devotee_type': request.POST.get('devotee_type', 'haribhakt'),
            'name': request.POST.get('name'),
            'contact_number': request.POST.get('contact_number'),
            'date_of_birth': request.POST.get('date_of_birth'),
            'gender': request.POST.get('gender'),
            'age': int(request.POST.get('age', 0)),
            'sabha_type': request.POST.get('sabha_type'),
            'address_line': request.POST.get('address_line'),
            'landmark': request.POST.get('landmark'),
            'zone': request.POST.get('zone'),
            'join_date': request.POST.get('join_date'),
            'photo_url': devotee.get('photo_url', ''),
            'updated_at': datetime.now().isoformat()
        }
        
        # Handle photo upload
        if request.POST.get('cropped_photo'):
            # Handle cropped photo from base64
            import base64
            from io import BytesIO
            from django.core.files.uploadedfile import InMemoryUploadedFile
            
            cropped_data = request.POST.get('cropped_photo')
            format, imgstr = cropped_data.split(';base64,')
            ext = format.split('/')[-1]
            
            data = base64.b64decode(imgstr)
            photo_file = InMemoryUploadedFile(
                BytesIO(data), None, f'cropped_photo.{ext}', f'image/{ext}', len(data), None
            )
            
            photo_url = upload_devotee_photo(photo_file, update_data)
            if photo_url:
                update_data['photo_url'] = photo_url
        elif 'photo' in request.FILES:
            # Handle regular photo upload
            photo_file = request.FILES['photo']
            photo_url = upload_devotee_photo(photo_file, update_data)
            if photo_url:
                update_data['photo_url'] = photo_url
        
        update_result = devotees_db.update_one({'_id': ObjectId(pk)}, update_data)
        
        if update_result is None:
            messages.error(request, 'Database connection error. Please try again.')
            return render(request, 'attendance/devotee_form.html', {'devotee': devotee, 'title': 'Edit Devotee', 'today': datetime.now().date().isoformat()})
        
        print(f"Database update result: {update_result.modified_count if update_result else 0} documents modified")
        print(f"Final photo_url in update_data: {update_data.get('photo_url')}")
        
        messages.success(request, 'Devotee updated successfully!')
        return redirect('devotee_detail', pk=pk)
    
    return render(request, 'attendance/devotee_form.html', {'devotee': devotee, 'title': 'Edit Devotee', 'today': datetime.now().date().isoformat()})

@login_required
def sabha_list(request):
    sabhas_raw = sabhas_db.find(sort=[('date', -1)])
    sabhas = []
    for sabha in sabhas_raw:
        sabha['id'] = str(sabha['_id'])
        sabha['get_sabha_type_display'] = sabha.get('sabha_type', '').title() + ' Sabha'
        sabhas.append(sabha)
    return render(request, 'attendance/sabha_list.html', {'sabhas': sabhas})

@login_required
def sabha_add(request):
    if request.method == 'POST':
        sabha_data = {
            'date': request.POST.get('date'),
            'sabha_type': request.POST.get('sabha_type'),
            'location': request.POST.get('location'),
            'xetra': request.POST.get('xetra'),
            'mandal': request.POST.get('mandal'),
            'start_time': request.POST.get('start_time'),
            'end_time': request.POST.get('end_time'),
            'created_at': datetime.now().isoformat()
        }
        result = sabhas_db.insert_one(sabha_data)
        sabha_id = str(result.inserted_id)
        
        # Create default absent records for all devotees of this sabha type
        all_devotees = devotees_db.find({'sabha_type': sabha_data['sabha_type']})
        for devotee in all_devotees:
            attendance_db.insert_one({
                'devotee_id': devotee.get('devotee_id', 'N/A'),
                'sabha_id': sabha_id,
                'status': 'absent',
                'notes': '',
                'timestamp': datetime.now().isoformat()
            })
        
        messages.success(request, 'Sabha created successfully!')
        return redirect('sabha_list')
    
    return render(request, 'attendance/sabha_form.html', {'title': 'Create Sabha', 'today': datetime.now().date().isoformat()})

@login_required
def mark_attendance(request, sabha_id):
    sabha = sabhas_db.find_one({'_id': ObjectId(sabha_id)})
    if not sabha:
        messages.error(request, 'Sabha not found')
        return redirect('sabha_list')
    sabha['id'] = str(sabha['_id'])
    sabha['get_sabha_type_display'] = sabha.get('sabha_type', '').title() + ' Sabha'
    
    search_query = request.GET.get('search', '')
    search_type = request.GET.get('search_type', 'id')
    page = int(request.GET.get('page', 1))
    per_page = 30
    
    query = {'sabha_type': sabha['sabha_type']}
    
    if search_query:
        search_condition = {}
        if search_type == 'id':
            # Search both as integer and string to handle mixed data types
            if search_query.isdigit():
                search_condition = {'$or': [
                    {'devotee_id': int(search_query)},
                    {'devotee_id': str(search_query)}
                ]}
            else:
                search_condition = {'devotee_id': {'$regex': str(search_query), '$options': 'i'}}
        elif search_type == 'phone':
            search_condition = {'contact_number': {'$regex': search_query}}
        elif search_type == 'name':
            search_condition = {'name': {'$regex': search_query, '$options': 'i'}}
        elif search_type == 'type':
            search_condition = {'devotee_type': {'$regex': search_query, '$options': 'i'}}
        
        query = {'$and': [{'sabha_type': sabha['sabha_type']}, search_condition]}
    
    total_count = devotees_db.count(query)
    skip = (page - 1) * per_page
    
    # Get all matching documents and sort by devotee_id as integer
    all_devotees = list(devotees_db.find(query))
    all_devotees.sort(key=lambda x: int(x.get('devotee_id', 0)) if str(x.get('devotee_id', '')).isdigit() else 0)
    devotees_raw = all_devotees[skip:skip + per_page]
    devotees = []
    for devotee in devotees_raw:
        devotee['id'] = str(devotee['_id'])
        devotee['get_sabha_type_display'] = devotee.get('sabha_type', '').title()
        if not devotee.get('photo_url'):
            devotee['photo_url'] = generate_initials_photo(devotee['name'])
        devotees.append(devotee)
    
    # Pagination info
    total_pages = (total_count + per_page - 1) // per_page
    has_previous = page > 1
    has_next = page < total_pages
    
    if request.method == 'POST':
        for devotee in devotees:
            devotee_object_id = str(devotee['_id'])
            status = request.POST.get(f'status_{devotee_object_id}', 'absent')
            notes = request.POST.get(f'notes_{devotee_object_id}', '')
            
            devotee_number_id = devotee.get('devotee_id', 'N/A')
            existing = attendance_db.find_one({
                'devotee_id': devotee_number_id,
                'sabha_id': sabha_id
            })
            
            if existing:
                attendance_db.update_one(
                    {'_id': existing['_id']},
                    {'status': status, 'notes': notes, 'timestamp': datetime.now().isoformat()}
                )
            else:
                attendance_db.insert_one({
                    'devotee_id': devotee_number_id,
                    'sabha_id': sabha_id,
                    'status': status,
                    'notes': notes,
                    'timestamp': datetime.now().isoformat()
                })
        
        messages.success(request, 'Attendance marked successfully!')
        # Redirect back to same page to maintain pagination
        redirect_url = f'/sabhas/{sabha_id}/attendance/?page={page}'
        if search_query:
            redirect_url += f'&search={search_query}'
        return redirect(redirect_url)
    
    # Get existing attendance for current page only
    devotee_number_ids = [d.get('devotee_id', 'N/A') for d in devotees]
    existing_attendance = {}
    for att in attendance_db.find({
        'sabha_id': sabha_id,
        'devotee_id': {'$in': devotee_number_ids}
    }):
        existing_attendance[att['devotee_id']] = att
    
    # Add default absent status for devotees without records
    for devotee in devotees:
        devotee_number_id = devotee.get('devotee_id', 'N/A')
        if devotee_number_id not in existing_attendance:
            existing_attendance[devotee_number_id] = {
                'status': 'absent',
                'notes': ''
            }
    

    
    # Create pagination object
    class PaginationObj:
        def __init__(self, items, page, total_pages, has_previous, has_next, total_count):
            self.object_list = items
            self.number = page
            self.paginator = type('Paginator', (), {
                'num_pages': total_pages,
                'count': total_count
            })()
            self.has_previous = lambda: has_previous
            self.has_next = lambda: has_next
            self.previous_page_number = lambda: page - 1 if has_previous else None
            self.next_page_number = lambda: page + 1 if has_next else None
        
        def __iter__(self):
            return iter(self.object_list)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        devotees_data = []
        for d in devotees_raw:
            photo_url = d.get('photo_url', '')
            if not photo_url:
                photo_url = generate_initials_photo(d['name'])
            devotees_data.append({
                'id': str(d['_id']),
                'devotee_id': d.get('devotee_id', ''),
                'name': d['name'],
                'contact_number': d['contact_number'],
                'sabha_type_display': d['sabha_type'].title(),
                'devotee_type': d.get('devotee_type', ''),
                'photo_url': photo_url
            })
        
        return JsonResponse({
            'devotees': devotees_data,
            'total_count': total_count,
            'current_page': page,
            'total_pages': total_pages,
            'has_previous': has_previous,
            'has_next': has_next
        })
    
    page_obj = PaginationObj(devotees, page, total_pages, has_previous, has_next, total_count)
    
    context = {
        'sabha': sabha,
        'devotees': devotees,
        'page_obj': page_obj,
        'existing_attendance': existing_attendance,
        'search_query': search_query,
        'total_count': total_count
    }
    return render(request, 'attendance/mark_attendance.html', context)

@login_required
def attendance_analytics(request):
    from collections import defaultdict
    from datetime import timedelta
    
    # Get filters
    sabha_type = request.GET.get('sabha_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Build query
    query = {}
    if sabha_type:
        query['sabha_type'] = sabha_type
    if date_from:
        query.setdefault('sabha_date', {})['$gte'] = date_from
    if date_to:
        query.setdefault('sabha_date', {})['$lte'] = date_to
    
    # Get all attendance records and join with sabha info
    records = attendance_db.find(query)
    
    # Aggregate data
    sabha_counts = defaultdict(int)
    weekly_data = defaultdict(lambda: defaultdict(int))
    monthly_trend = defaultdict(int)
    date_intensity = defaultdict(int)
    
    for record in records:
        if record['status'] == 'present':
            # Get sabha info
            sabha = sabhas_db.find_one({'_id': ObjectId(record['sabha_id'])})
            if sabha:
                sabha_counts[sabha['sabha_type']] += 1
                date_intensity[sabha['date']] += 1
                
                # Monthly trend (last 3 months)
                try:
                    date_obj = datetime.fromisoformat(sabha['date'])
                    month_key = date_obj.strftime('%Y-%m')
                    monthly_trend[month_key] += 1
                except:
                    pass
    
    # Prepare response
    data = {
        'sabha_counts': dict(sabha_counts),
        'monthly_trend': dict(sorted(monthly_trend.items())),
        'date_intensity': dict(date_intensity),
        'total': sum(sabha_counts.values())
    }
    
    return JsonResponse(data)

@login_required
def attendance_report(request):
    sabha_type = request.GET.get('sabha_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    per_page = 20
    
    # Build sabha query first
    sabha_query = {}
    if sabha_type:
        sabha_query['sabha_type'] = sabha_type
    if date_from:
        sabha_query.setdefault('date', {})['$gte'] = date_from
    if date_to:
        sabha_query.setdefault('date', {})['$lte'] = date_to
    
    # Get matching sabha IDs
    matching_sabhas = sabhas_db.find(sabha_query, {'_id': 1})
    sabha_ids = [str(s['_id']) for s in matching_sabhas]
    
    # Build attendance query
    query = {}
    if sabha_ids:
        query['sabha_id'] = {'$in': sabha_ids}
    if status_filter:
        query['status'] = status_filter
    search_type = request.GET.get('search_type', 'name')
    if search_query:
        if search_type == 'name':
            # Get devotee IDs that match name search
            matching_devotees = devotees_db.find(
                {'name': {'$regex': search_query, '$options': 'i'}}, 
                {'devotee_id': 1}
            )
            devotee_ids = [d.get('devotee_id') for d in matching_devotees if d.get('devotee_id')]
            if devotee_ids:
                query['devotee_id'] = {'$in': devotee_ids}
            else:
                query['devotee_id'] = {'$in': []}  # No matches
        elif search_type == 'id':
            # Search directly by devotee_id in attendance records
            if search_query.isdigit():
                query['devotee_id'] = int(search_query)
            else:
                query['devotee_id'] = {'$regex': str(search_query), '$options': 'i'}
        elif search_type == 'phone':
            # Get devotee IDs that match phone search, then search attendance by those IDs
            matching_devotees = devotees_db.find(
                {'contact_number': {'$regex': search_query}}, 
                {'devotee_id': 1}
            )
            devotee_ids = [d.get('devotee_id') for d in matching_devotees if d.get('devotee_id')]
            if devotee_ids:
                query['devotee_id'] = {'$in': devotee_ids}
            else:
                query['devotee_id'] = {'$in': []}  # No matches
    
    # Get counts for status buttons
    base_query = {k: v for k, v in query.items() if k != 'status'}
    present_count = attendance_db.count({**base_query, 'status': 'present'})
    absent_count = attendance_db.count({**base_query, 'status': 'absent'})
    late_count = attendance_db.count({**base_query, 'status': 'late'})
    
    # Pagination
    total_count = attendance_db.count(query)
    skip = (page - 1) * per_page
    attendance_records = attendance_db.find(query, sort=[('sabha_date', -1)], skip=skip, limit=per_page)
    
    # Enrich attendance records with devotee and sabha info
    attendance_list = list(attendance_records)
    for record in attendance_list:
        # Get devotee info
        devotee = devotees_db.find_one({'devotee_id': record['devotee_id']})
        if devotee:
            record['devotee_name'] = devotee['name']
            record['mobile'] = devotee.get('contact_number', 'N/A')
            record['devotee_type'] = devotee.get('devotee_type', 'N/A')
        else:
            record['devotee_name'] = 'Unknown'
            record['mobile'] = 'N/A'
            record['devotee_type'] = 'N/A'
        
        # Get sabha info
        sabha = sabhas_db.find_one({'_id': ObjectId(record['sabha_id'])})
        if sabha:
            record['sabha_date'] = sabha['date']
            record['sabha_type'] = sabha['sabha_type']
        else:
            record['sabha_date'] = 'Unknown'
            record['sabha_type'] = 'Unknown'
    
    # Pagination info
    total_pages = (total_count + per_page - 1) // per_page
    has_previous = page > 1
    has_next = page < total_pages
    
    # Create pagination object
    class PaginationObj:
        def __init__(self, items, page, total_pages, has_previous, has_next, total_count):
            self.object_list = items
            self.number = page
            self.paginator = type('Paginator', (), {
                'num_pages': total_pages,
                'count': total_count
            })()
            self.has_previous = lambda: has_previous
            self.has_next = lambda: has_next
            self.previous_page_number = lambda: page - 1 if has_previous else None
            self.next_page_number = lambda: page + 1 if has_next else None
        
        def __iter__(self):
            return iter(self.object_list)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Convert ObjectId to string for JSON serialization
        serializable_records = []
        for record in attendance_list:
            clean_record = {}
            for key, value in record.items():
                if isinstance(value, ObjectId):
                    clean_record[key] = str(value)
                else:
                    clean_record[key] = value
            serializable_records.append(clean_record)
        
        return JsonResponse({
            'records': serializable_records,
            'total_count': total_count,
            'current_page': page,
            'total_pages': total_pages,
            'has_previous': has_previous,
            'has_next': has_next,
            'present_count': present_count,
            'absent_count': absent_count,
            'late_count': late_count
        })
    
    # Convert ObjectId to string for template rendering
    for record in attendance_list:
        if '_id' in record:
            record['_id'] = str(record['_id'])
    
    page_obj = PaginationObj(attendance_list, page, total_pages, has_previous, has_next, total_count)
    
    context = {
        'page_obj': page_obj,
        'sabha_types': [('bal', 'Bal Sabha'), ('yuvak', 'Yuvak Sabha'), ('mahila', 'Mahila Sabha'), ('sanyukt', 'Sanyukt Sabha')],
        'filters': {'sabha_type': sabha_type, 'date_from': date_from, 'date_to': date_to, 'status': status_filter},
        'search_query': search_query,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'total_count': total_count
    }
    return render(request, 'attendance/attendance_report.html', context)

@login_required
def export_attendance(request):
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from concurrent.futures import ThreadPoolExecutor
    from io import BytesIO
    
    sabha_type_filter = request.GET.get('sabha_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status_filter = request.GET.get('status')
    
    # Build queries
    sabha_query = {}
    if sabha_type_filter:
        sabha_query['sabha_type'] = sabha_type_filter
    if date_from:
        sabha_query.setdefault('date', {})['$gte'] = date_from
    if date_to:
        sabha_query.setdefault('date', {})['$lte'] = date_to
    
    # Get data using pandas for fast processing
    sabhas = list(sabhas_db.find(sabha_query, sort=[('date', 1)]))
    sabha_ids = [str(s['_id']) for s in sabhas]
    
    att_query = {'sabha_id': {'$in': sabha_ids}} if sabha_ids else {}
    if status_filter:
        att_query['status'] = status_filter
    
    attendance_records = list(attendance_db.find(att_query))
    devotee_ids = list(set(att['devotee_id'] for att in attendance_records))
    devotees = list(devotees_db.find({'devotee_id': {'$in': devotee_ids}}))
    
    # Create pandas DataFrames for fast processing
    sabha_df = pd.DataFrame(sabhas)
    devotee_df = pd.DataFrame(devotees)
    att_df = pd.DataFrame(attendance_records)
    
    if not att_df.empty:
        # Convert ObjectId to string for merge
        sabha_df['_id_str'] = sabha_df['_id'].astype(str)
        
        # Merge data using pandas
        att_df = att_df.merge(sabha_df[['_id_str', 'sabha_type', 'date']], left_on='sabha_id', right_on='_id_str', how='left')
        att_df = att_df.merge(devotee_df[['devotee_id', 'name', 'contact_number', 'devotee_type']], on='devotee_id', how='left')
        
        # Debug merged data
        print(f"Merged data columns: {att_df.columns.tolist()}")
        print(f"Sample sabha_type values: {att_df['sabha_type'].unique()}")
        
        # Filter out rows with missing sabha_type
        att_df = att_df.dropna(subset=['sabha_type'])
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    
    def create_summary_sheet():
        ws = wb.active
        ws.title = '📊 Summary'
        
        if att_df.empty:
            return
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = '🏛️ Temple Attendance Summary by Date'
        title_cell.font = Font(bold=True, size=16, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
        title_cell.alignment = center_align
        
        # Date-wise summary for all sabha types
        date_summary = att_df.groupby(['date', 'sabha_type', 'status']).size().unstack(fill_value=0).reset_index()
        
        headers = ['Date', 'Sabha Type', 'Present', 'Absent', 'Late']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        row = 4
        for _, data in date_summary.iterrows():
            ws.cell(row=row, column=1, value=str(data['date']))
            ws.cell(row=row, column=2, value=data['sabha_type'].title())
            ws.cell(row=row, column=3, value=data.get('present', 0))
            ws.cell(row=row, column=4, value=data.get('absent', 0))
            ws.cell(row=row, column=5, value=data.get('late', 0))
            row += 1
    
    def create_sabha_sheet(sabha_type):
        sabha_data = att_df[att_df['sabha_type'] == sabha_type]
        
        if sabha_data.empty:
            return None
        
        try:
            ws = wb.create_sheet(f'📋 {sabha_type.title()} Sabha')
            
            # Create pivot table with devotees as rows and dates as columns
            pivot_data = sabha_data.pivot_table(
                index=['devotee_id', 'name', 'contact_number', 'devotee_type'],
                columns='date',
                values='status',
                aggfunc='first',
                fill_value='absent'
            ).reset_index()
            
            # Get unique dates for column count
            dates = sorted(sabha_data['date'].unique())
            
            # Title
            title_end_col = chr(65 + len(pivot_data.columns) - 1)
            ws.merge_cells(f'A1:{title_end_col}1')
            ws['A1'] = f'🏛️ {sabha_type.title()} Sabha - Devotee Attendance by Date'
            ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
            ws['A1'].alignment = center_align
            
            # Date-wise counts above each date column
            for col_idx, date in enumerate(dates, 5):
                date_data = sabha_data[sabha_data['date'] == date]
                present_count = len(date_data[date_data['status'] == 'present'])
                absent_count = len(date_data[date_data['status'] == 'absent'])
                late_count = len(date_data[date_data['status'] == 'late'])
                
                count_text = f'✅{present_count} ❌{absent_count} ⏰{late_count}'
                cell = ws.cell(row=2, column=col_idx, value=count_text)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                cell.alignment = center_align
            
            # Headers
            headers = ['ID', 'Name', 'Mobile', 'Type'] + [str(date) for date in dates]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # Data rows
            for idx, row in pivot_data.iterrows():
                excel_row = idx + 4
                ws.cell(row=excel_row, column=1, value=row['devotee_id'])
                ws.cell(row=excel_row, column=2, value=row['name'])
                ws.cell(row=excel_row, column=3, value=row['contact_number'])
                ws.cell(row=excel_row, column=4, value=row['devotee_type'])
                
                # Fill date columns with status and color coding
                for col_idx, date in enumerate(dates, 5):
                    status = row[date] if date in row else 'absent'
                    cell = ws.cell(row=excel_row, column=col_idx, value=status.title())
                    
                    # Color coding
                    if status == 'present':
                        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    elif status == 'late':
                        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            # Add summary row at bottom
            summary_row = len(pivot_data) + 6
            ws.cell(row=summary_row, column=1, value='TOTALS:').font = Font(bold=True)
            
            for col_idx, date in enumerate(dates, 5):
                date_data = sabha_data[sabha_data['date'] == date]
                present_count = len(date_data[date_data['status'] == 'present'])
                absent_count = len(date_data[date_data['status'] == 'absent'])
                late_count = len(date_data[date_data['status'] == 'late'])
                
                summary_text = f'✅{present_count} ❌{absent_count} ⏰{late_count}'
                cell = ws.cell(row=summary_row, column=col_idx, value=summary_text)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            return ws
            
        except Exception as e:
            print(f"Error creating sheet for {sabha_type}: {str(e)}")
            return None
    
    # Create summary sheet
    create_summary_sheet()
    
    # Create sabha sheets
    if not att_df.empty:
        sabha_types = att_df['sabha_type'].unique()
        print(f"Found sabha types: {sabha_types}")  # Debug
        for sabha_type in sabha_types:
            print(f"Creating sheet for: {sabha_type}")  # Debug
            sheet = create_sabha_sheet(sabha_type)
            if sheet:
                print(f"Successfully created sheet: {sheet.title}")  # Debug
            else:
                print(f"Failed to create sheet for: {sabha_type}")  # Debug

    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="temple_attendance_report.xlsx"'
    
    return response

@login_required
def upload_devotees(request):
    from .forms import DevoteeUploadForm
    from .utils import process_excel_file
    import tempfile
    import os
    import gc
    
    if request.method == 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return process_devotees_batch(request)
        
        # Check if upload is already in progress
        if 'upload_data' in request.session:
            messages.warning(request, 'Upload already in progress. Please wait for completion.')
            context = {
                'form': DevoteeUploadForm(),
                'show_progress': True,
                'total_records': request.session.get('total_records', 0)
            }
            return render(request, 'attendance/upload_devotees.html', context)
        
        form = DevoteeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            sabha_type_filter = form.cleaned_data.get('sabha_type_filter')
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in excel_file.chunks():
                    tmp_file.write(chunk)
                tmp_file_path = tmp_file.name
            
            try:
                result, error = process_excel_file(tmp_file_path, sabha_type_filter)
                
                if error:
                    messages.error(request, error)
                    return render(request, 'attendance/upload_devotees.html', {'form': form})
                
                valid_rows = result['valid_rows']
                errors = result['errors']
                
                if errors:
                    context = {
                        'form': form,
                        'errors': errors,
                        'total_rows': len(valid_rows) + len(errors),
                        'valid_count': len(valid_rows),
                        'error_count': len(errors)
                    }
                    return render(request, 'attendance/upload_devotees.html', context)
                
                # Convert dates to strings for JSON serialization
                for row in valid_rows:
                    if 'join_date' in row and hasattr(row['join_date'], 'isoformat'):
                        row['join_date'] = row['join_date'].isoformat()
                    if 'date_of_birth' in row and hasattr(row['date_of_birth'], 'isoformat'):
                        row['date_of_birth'] = row['date_of_birth'].isoformat()
                
                request.session['upload_data'] = valid_rows
                request.session['total_records'] = len(valid_rows)
                request.session['current_batch'] = 0
                request.session['total_created'] = 0
                request.session['total_updated'] = 0
                
                context = {
                    'form': form,
                    'show_progress': True,
                    'total_records': len(valid_rows)
                }
                return render(request, 'attendance/upload_devotees.html', context)
                
            except Exception as e:
                messages.error(request, f'Error during import: {str(e)}')
                return render(request, 'attendance/upload_devotees.html', {'form': form})
            finally:
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)
                gc.collect()
    else:
        # Check if upload is in progress on GET request
        if 'upload_data' in request.session:
            context = {
                'form': DevoteeUploadForm(),
                'show_progress': True,
                'total_records': request.session.get('total_records', 0)
            }
            return render(request, 'attendance/upload_devotees.html', context)
        
        form = DevoteeUploadForm()
    
    return render(request, 'attendance/upload_devotees.html', {'form': form})

@login_required
def cancel_batch_processing(request):
    # Clear all batch processing session data
    keys_to_remove = ['upload_data', 'total_records', 'current_batch', 'total_created', 'total_updated']
    for key in keys_to_remove:
        if key in request.session:
            del request.session[key]
    return JsonResponse({'success': True})

def process_batch_worker(batch_data):
    """Worker function to process a batch of devotees"""
    from .mongodb_utils import MongoDBManager
    
    devotees_worker_db = MongoDBManager('devotees')
    created_count = 0
    updated_count = 0
    
    for row in batch_data:
        try:
            # Clean row data
            clean_row = {}
            for k, v in row.items():
                if isinstance(v, ObjectId):
                    clean_row[k] = str(v)
                elif hasattr(v, 'isoformat'):
                    clean_row[k] = v.isoformat()
                else:
                    clean_row[k] = v
            
            # Check only by devotee_id for uniqueness
            existing = devotees_worker_db.find_one({'devotee_id': clean_row.get('devotee_id')})
            if existing:
                # Overwrite existing devotee with same devotee_id
                update_data = {k: v for k, v in clean_row.items() if k != '_id'}
                update_data['updated_at'] = datetime.now().isoformat()
                devotees_worker_db.update_one({'_id': existing['_id']}, update_data)
                updated_count += 1
            else:
                # Create new devotee (even if contact_number exists elsewhere)
                clean_row['created_at'] = datetime.now().isoformat()
                devotees_worker_db.collection.insert_one(clean_row)
                created_count += 1
        except Exception as e:
            print(f"Error processing row: {e}")
            continue
    
    return created_count, updated_count

@login_required
def process_devotees_batch(request):
    if 'upload_data' not in request.session:
        return JsonResponse({'error': 'No upload data found'})
    
    from concurrent.futures import ThreadPoolExecutor
    import threading
    
    batch_size = 25  # Smaller batches for parallel processing
    num_threads = 4  # Process 4 batches in parallel
    current_batch = request.session.get('current_batch', 0)
    
    valid_rows = request.session['upload_data']
    total_records = len(valid_rows)
    
    # Calculate batch range for parallel processing
    start_idx = current_batch * batch_size * num_threads
    end_idx = min(start_idx + (batch_size * num_threads), total_records)
    
    if start_idx >= total_records:
        return JsonResponse({
            'processed': total_records,
            'total': total_records,
            'created': request.session.get('total_created', 0),
            'updated': request.session.get('total_updated', 0),
            'complete': True,
            'percentage': 100
        })
    
    # Split data into parallel batches
    batches = []
    for i in range(num_threads):
        batch_start = start_idx + (i * batch_size)
        batch_end = min(batch_start + batch_size, end_idx)
        if batch_start < batch_end:
            batches.append(valid_rows[batch_start:batch_end])
    
    total_created = 0
    total_updated = 0
    
    # Process batches in parallel
    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = [executor.submit(process_batch_worker, batch) for batch in batches]
        
        for future in futures:
            try:
                created, updated = future.result(timeout=30)
                total_created += created
                total_updated += updated
            except Exception as e:
                print(f"Batch processing error: {e}")
    
    # Update session counters
    request.session['current_batch'] = current_batch + 1
    request.session['total_created'] = request.session.get('total_created', 0) + total_created
    request.session['total_updated'] = request.session.get('total_updated', 0) + total_updated
    
    processed = end_idx
    is_complete = processed >= total_records
    
    if is_complete:
        for key in ['upload_data', 'total_records', 'current_batch', 'total_created', 'total_updated']:
            request.session.pop(key, None)
    
    return JsonResponse({
        'processed': processed,
        'total': total_records,
        'created': request.session.get('total_created', 0),
        'updated': request.session.get('total_updated', 0),
        'complete': is_complete,
        'percentage': round((processed / total_records) * 100),
        'current_batch': current_batch + 1
    })

@login_required
def devotee_delete(request, pk):
    devotee = devotees_db.find_one({'_id': ObjectId(pk)})
    if not devotee:
        messages.error(request, 'Devotee not found')
        return redirect('devotee_list')
    
    if request.method == 'POST':
        # Delete devotee and related attendance records
        attendance_db.delete_many({'devotee_id': pk})
        devotees_db.delete_one({'_id': ObjectId(pk)})
        messages.success(request, f'Devotee {devotee["name"]} deleted successfully!')
        return redirect('devotee_list')
    
    return JsonResponse({'name': devotee['name']})

@login_required
def sabha_delete(request, pk):
    sabha = sabhas_db.find_one({'_id': ObjectId(pk)})
    if not sabha:
        messages.error(request, 'Sabha not found')
        return redirect('sabha_list')
    
    if request.method == 'POST':
        # Delete sabha and related attendance records
        attendance_db.delete_many({'sabha_id': pk})
        sabhas_db.delete_one({'_id': ObjectId(pk)})
        messages.success(request, f'Sabha on {sabha["date"]} deleted successfully!')
        return redirect('sabha_list')
    
    return JsonResponse({'date': sabha['date'], 'type': sabha['sabha_type']})

@login_required
@require_POST
def save_individual_attendance(request):
    try:
        data = json.loads(request.body)
        sabha_id = data.get('sabha_id')
        devotee_id = data.get('devotee_id')
        status = data.get('status', 'absent')
        notes = data.get('notes', '')
        
        sabha = sabhas_db.find_one({'_id': ObjectId(sabha_id)})
        devotee = devotees_db.find_one({'_id': ObjectId(devotee_id)})
        
        if not sabha or not devotee:
            return JsonResponse({'success': False, 'error': 'Sabha or Devotee not found'})
        
        devotee = devotees_db.find_one({'_id': ObjectId(devotee_id)})
        if not devotee:
            return JsonResponse({'success': False, 'error': 'Devotee not found'})
            
        devotee_number_id = devotee.get('devotee_id', 'N/A')
        existing = attendance_db.find_one({'devotee_id': devotee_number_id, 'sabha_id': sabha_id})
        
        if existing:
            attendance_db.update_one(
                {'_id': existing['_id']},
                {'status': status, 'notes': notes, 'timestamp': datetime.now().isoformat()}
            )
        else:
            attendance_db.insert_one({
                'devotee_id': devotee_number_id,
                'sabha_id': sabha_id,
                'status': status,
                'notes': notes,
                'timestamp': datetime.now().isoformat()
            })
        
        return JsonResponse({'success': True, 'message': 'Attendance saved'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})
